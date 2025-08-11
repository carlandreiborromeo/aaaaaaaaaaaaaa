from flask import Blueprint, request, send_file, jsonify
from openpyxl import load_workbook
import os
import tempfile

bp = Blueprint('generate', __name__, url_prefix='/api/generate')

def to_number(val):
    """Convert to int/float if numeric, else return original or None."""
    try:
        if val is None or str(val).strip() == "":
            return None
        num = float(val)
        return int(num) if num.is_integer() else num
    except (ValueError, TypeError):
        return val  # Keep text as-is

@bp.route('/excel', methods=['POST'])
def generate_excel():
    try:
        students = request.json.get("students", [])
        if not students:
            return jsonify({"error": "No student data received"}), 400

        # Load template
        template_path = os.path.join(os.path.dirname(__file__), '..', 'templates', 'Grades.xlsx')
        wb = load_workbook(template_path)


        # Map departments to worksheets
        sheet_map = {
            "PRODUCTION": wb["PRODUCTION"],
            "SUPPORT": wb["SUPPORT"],
            "TECHNICAL": wb["TECHNICAL"]
        }

        # Get first student's values for headers
        first_student = students[0] if students else {}
        immersion_date = first_student.get("date_of_immersion", "")
        batch = first_student.get("batch", "")
        school = first_student.get("school", "")

        # Fill header cells for all sheets
        for ws in wb.worksheets:
            ws['H8'] = batch + " - " + school
            ws['H9'] = immersion_date

        # Row counters for each sheet
        row_counter = { "PRODUCTION": 10, "SUPPORT": 10, "TECHNICAL": 10 }

        for s in students:
            dept_raw = (s.get("department") or "").strip().upper()
            if dept_raw in ["TECHNICAL", "IT"]:
                dept = "TECHNICAL"
            elif dept_raw == "PRODUCTION":
                dept = "PRODUCTION"
            else:
                dept = "SUPPORT"

            ws = sheet_map[dept]
            row = row_counter[dept]

            # Fill basic info
            ws[f'B{row}'] = s.get("last_name", "")
            ws[f'C{row}'] = s.get("first_name", "")
            ws[f'D{row}'] = s.get("middle_name", "")
            ws[f'E{row}'] = s.get("strand", "")
            ws[f'F{row}'] = s.get("department", "")
            ws[f'G{row}'] = to_number(s.get("over_all", ""))  # Performance Appraisal

            # Fill grades with numeric-safe conversion
            ws[f'G{row}'] = to_number(s.get("1G", ""))
            ws[f'H{row}'] = to_number(s.get("1G", ""))
            ws[f'I{row}'] = to_number(s.get("2G", ""))
            ws[f'J{row}'] = to_number(s.get("3G", ""))
            ws[f'K{row}'] = to_number(s.get("4G", ""))
            ws[f'L{row}'] = to_number(s.get("5G", ""))
            ws[f'M{row}'] = to_number(s.get("6G", ""))
            ws[f'N{row}'] = to_number(s.get("7G", ""))
            ws[f'O{row}'] = to_number(s.get("8G", ""))
            ws[f'P{row}'] = to_number(s.get("9G", ""))
            ws[f'Q{row}'] = to_number(s.get("10G", ""))
            ws[f'R{row}'] = to_number(s.get("11G", ""))
            ws[f'S{row}'] = to_number(s.get("12G", ""))

            if dept == "PRODUCTION":
                ws[f'V{row}'] = to_number(s.get("13G", ""))
                ws[f'W{row}'] = to_number(s.get("14G", ""))
                ws[f'X{row}'] = to_number(s.get("15G", ""))
                ws[f'Y{row}'] = to_number(s.get("16G", ""))
                ws[f'AB{row}'] = to_number(s.get("17G", ""))           
                ws[f'AC{row}'] = to_number(s.get("18G", ""))

            if dept == "SUPPORT":
                ws[f'U{row}'] = to_number(s.get("13G", ""))
                ws[f'Z{row}'] = to_number(s.get("14G", ""))
                ws[f'AC{row}'] = to_number(s.get("15G", ""))
            # Only Technical gets the 15th grade
            if dept == "TECHNICAL":
                ws[f'T{row}'] = to_number(s.get("13G", ""))
                ws[f'AA{row}'] = to_number(s.get("14G", ""))
                ws[f'AC{row}'] = to_number(s.get("15G", ""))

            row_counter[dept] += 1

        # Save temp file
        temp_dir = tempfile.mkdtemp()
        output_path = os.path.join(temp_dir, "generated_immersion_report.xlsx")
        wb.save(output_path)

        return send_file(output_path, as_attachment=True)

    except Exception as e:
        return jsonify({"error": str(e)}), 500
